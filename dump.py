from collections import OrderedDict
import xlsxwriter
from openpyxl import Workbook, load_workbook

def unpack(s, t=None):
    if t is None:
        t = str(s)[2:]
        s = str(s)[0:2]
    s = int(s, 16)
    t = int(t, 16)
    value = (t * 0x100) + s
    return value

def pack(h):
    s = h % 0x100
    t = h // 0x100
    return (s, t)

def ascii_to_hex_string(eng, control_codes={}):
    """Returns a hex string of the ascii bytes of a given english (translated) string."""
    eng_bytestring = ""
    if not eng:
        return ""
    else:
        try:
            eng = str(eng)
        except UnicodeEncodeError:
            # Tried to encode a fullwidth number. Encode it as sjis instead.
            eng = eng.encode('shift-jis')

        eng_bytestring = eng

        for cc in control_codes:
            cc_hex = ascii_to_hex_string(cc)
            if cc_hex in eng_bytestring:
                eng_bytestring = eng_bytestring.replace(cc_hex, control_codes[cc])

        return eng_bytestring

def sjis_to_hex_string(jp, control_codes={}):
    """Returns a hex string of the Shift JIS bytes of a given japanese string."""
    jp_bytestring = ""
    try:
        sjis = jp.encode('shift-jis')
    except AttributeError:
        # Trying to encode numbers throws an attribute error; they aren't important, so just keep the number
        sjis = str(jp)
        
    jp_bytestring = sjis

    for cc in control_codes:
        cc_hex = sjis_to_hex_string(cc)
        if cc_hex in jp_bytestring:
            jp_bytestring = jp_bytestring.replace(cc_hex, control_codes[cc])

    return jp_bytestring


class Translation(object):
    """Has an offset, a SJIS japanese string, and an ASCII english string."""
    def __init__(self, gamefile, location, japanese, english):
        self.location = location
        self.gamefile = gamefile
        self.japanese = japanese
        self.english = english

        #self.jp_bytestring = sjis_to_hex_string(japanese)
        #self.en_bytestring = ascii_to_hex_string(english)
        self.jp_bytestring = japanese
        self.en_bytestring = english

    def refresh_jp_bytestring(self):
        self.jp_bytestring = sjis_to_hex_string(self.japanese)
        return self.jp_bytestring

    def refresh_en_bytestring(self):
        self.en_bytestring = ascii_to_hex_string(self.english)
        return self.en_bytestring

    def __repr__(self):
        return "%s %s" % (hex(self.location), self.english)


class BorlandPointer(object):
    """Two-byte, little-endian pointer with a constant added to retrieve location."""
    def __init__(self, gamefile, pointer_location, text_location):
        # A BorlandPointer has only one location. The container OrderDicts have lists of pointers,
        # each having their own location.
        self.gamefile = gamefile
        self.constant = self.gamefile.pointer_constant
        self.location = pointer_location
        self.text_location = text_location
        self.new_text_location = text_location

    def text(self):
        gamefile_slice = self.gamefile.filestring[self.text_location:self.text_location+30]
        gamefile_slice = gamefile_slice.split('\x00')[0]
        try:
            gamefile_slice = gamefile_slice.decode('shift_jis')
        except:
            gamefile_slice = "weird bytes"
        return gamefile_slice

    def original_text(self):
        gamefile_slice = self.gamefile.original_filestring[self.text_location:self.text_location+45]
        gamefile_slice = gamefile_slice.split('\x00')[0]
        try:
            gamefile_slice = gamefile_slice.decode('shift_jis')
        except:
            gamefile_slice = "weird bytes"
        return gamefile_slice

    def move_pointer_location(self, diff):
        self.location += diff
        return self.location


    def edit(self, diff):
        #print hex(self.location)
        first = hex(ord(self.gamefile.filestring[self.location]))
        second = hex(ord(self.gamefile.filestring[self.location+1]))
        #print first, second
        old_value = unpack(first, second)
        new_value = old_value + diff

        new_bytes = pack(new_value)
        new_first, new_second = chr(new_bytes[0]), chr(new_bytes[1])
        self.gamefile.filestring = self.gamefile.filestring[:self.location] + new_first + new_second + self.gamefile.filestring[self.location+2:]
        self.new_text_location = new_value
        return new_first, new_second

    def __repr__(self):
        return "%s pointing to %s" % (hex(self.location), hex(self.new_text_location))



class DumpExcel(object):
    """
    Takes a dump excel path, and lets you get a block's translations from it.
    """
    # TODO: Currently uses the order of the excel sheet. Might want to sort it by text_location in get_translations()...

    def __init__(self, path, control_codes={}):
        self.path = path
        self.workbook = load_workbook(self.path)
        self.control_codes = control_codes

    def get_translations(self, target):
        """Get the translations for a file."""
        # Accepts a block or a gamefile as "target".
        trans = []    # translations[offset] = Translation()
        try:
            worksheet = self.workbook.get_sheet_by_name(target.gamefile.filename)
        except AttributeError:
            try:
                worksheet = self.workbook.get_sheet_by_name(target.filename)
            except KeyError:
                worksheet = self.workbook.get_sheet_by_name(target.filename.lstrip('decompressed_'))
        for row in worksheet.rows[1:]:  # Skip the first row, it's just labels
            try:
                offset = int(row[0].value, 16)
            except TypeError:
                # Either a blank line or a total value. Ignore it.
                break
            try:
                start, stop = target.start, target.stop
                if not (target.start <= offset < target.stop):
                    continue
            except AttributeError:
                pass

            if row[3].value is None:
                continue

            japanese = row[1].value.encode('shift-jis')
            english = row[3].value.encode('shift-jis')

            #if isinstance(japanese, long):
            #    # Causes some encoding problems? Trying to skip them for now
            #    continue

            # Yeah this is important - blank strings are None (non-iterable), so use "" instead.
            if not english:
                english = ""

            trans.append(Translation(target, offset, japanese, english))
        return trans

class PointerExcel(object):
    def __init__(self, path):
        self.path = path
        try:
            self.workbook = load_workbook(self.path)
        except IOError:
            self.workbook = xlsxwriter.Workbook(self.path)         

    def add_worksheet(self, title):
        self.worksheet = self.workbook.add_worksheet(title)
        header = self.workbook.add_format({'bold': True, 'align': 'center', 'bottom': True, 'bg_color': 'gray'})
        self.worksheet.write(0, 0, 'Text Loc', header)
        self.worksheet.write(0, 1, 'Ptr Loc', header)
        self.worksheet.write(0, 2, 'Points To', header)
        self.worksheet.write(0, 3, 'Comments', header)
        self.worksheet.set_column('A:A', 7)
        self.worksheet.set_column('B:B', 7)
        self.worksheet.set_column('C:C', 30)
        self.worksheet.set_column('D:D', 30)
        return self.worksheet

    def get_pointers(self, gamefile):
        pointers = OrderedDict()
        try:
            ws = self.workbook[gamefile.filename]
        except KeyError:
            try:
                ws = self.workbook[gamefile.filename.lstrip('decompressed_')]
            except KeyError:
                # Not in the worksheet; return an emtpty OrderedDict
                return pointers
        for i, row in enumerate(ws):
            if i == 0:
                continue
            text_location = int(row[0].value, 16)
            pointer_location = int(row[1].value, 16)
            ptr = BorlandPointer(gamefile, pointer_location, text_location)
            if text_location in pointers:
                pointers[text_location].append(ptr)
            else:
                pointers[text_location] = [ptr,]
        return pointers

    def close(self):
        self.workbook.close()